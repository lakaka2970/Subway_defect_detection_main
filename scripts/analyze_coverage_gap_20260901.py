# -*- coding: utf-8 -*-
"""
真实线路需求 vs 现有 16 大类：覆盖率精确量化（2026-09-01）

目的：回答一个此前只有粗略估计（~30%）的关键问题——
      「现有 16 大类到底覆盖了真实线路多少需求？」

方法：
  1. 读滨海快线真实缺陷登记 1013 条（器材 N / 零部件 O / 描述 P）
  2. 状态抽取：状态 = 描述.replace(器材,"").replace(零部件,"")
     （登记表描述严格是「器材+零部件+状态」的拼接，此法准确）
  3. 器材层：把每条的器材映射到现有 16 大类的器材词，或判为「未覆盖」
  4. 状态层：把每条的状态映射到现有 16 大类的状态原语，或判为「未覆盖」
  5. 输出按样本量加权的覆盖率 + 未覆盖清单（按样本量降序）

产出：
  - 控制台报告
  - docs/plans/9.01全量数据盘点/coverage_gap_analysis.json
  - docs/plans/9.01全量数据盘点/器材映射待确认模板.xlsx（给甲方业务专家填充）
"""
import sys, json, collections
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
XLSX = Path(r"D:/BaiduNetdiskDownload/缺陷表-滨海快线/data（1）.xlsx")
OUT_DIR = ROOT / "docs/plans/9.01全量数据盘点"
OUT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------- 现有 16 大类的器材词（规范 taxonomy 命名） ----------------
# 来源 docs/模型输出与缺陷类型对照表.md 第一节
LEGACY_PARTS = [
    ("VHBNM/VHBNL",  "垂直悬吊安装底座",   ["垂直悬吊安装底座"]),
    ("SVHBNM/SVHBNL", "垂直悬吊槽钢底座",  ["垂直悬吊槽钢底座"]),
    ("SVHTNL",       "垂直悬吊槽钢上方",   ["垂直悬吊槽钢上方"]),
    ("RHTBNM/RHTBNL", "刚性悬挂吊柱底座",  ["刚性悬挂吊柱", "吊柱"]),
    ("GWCSBNM/NL",   "地线线夹托板安装底座", ["地线线夹托板"]),
    ("GWCNM/GWCNL",  "地线线夹",           ["地线线夹", "地线卡子"]),
    ("BSBM",         "汇流排中间接头",     ["汇流排"]),
    ("INSD",         "绝缘子",             ["绝缘子"]),
    ("CBHPM/CBVPM",  "腕臂底座",           ["腕臂底座"]),
    ("DRPS",         "吊弦",               ["吊弦"]),
]

# ---------------- 现有 16 大类的状态原语 ----------------
LEGACY_STATES = [
    ("螺母缺失", ["螺母缺失", "螺帽缺失", "副螺帽缺失"]),
    ("螺母松动", ["螺母松动", "螺帽松动", "副螺帽松动"]),
    ("螺栓缺失", ["螺栓缺失"]),
    ("销钉缺失", ["销钉缺失", "开口销缺失"]),
    ("破损",     ["破损"]),
    ("不受力",   ["不受力"]),
]


def match_legacy_part(part_name: str):
    """真实器材名 -> 现有 16 大类的器材词（返回 (大类编码, 匹配词) 或 (None, None)）"""
    for code, canon, kws in LEGACY_PARTS:
        for kw in kws:
            if kw in part_name:
                return code, canon
    return None, None


def match_legacy_state(state: str):
    """真实状态 -> 现有 16 大类的状态原语"""
    for canon, kws in LEGACY_STATES:
        for kw in kws:
            if kw in state:
                return canon
    return None


def main():
    try:
        import openpyxl
    except ImportError:
        print("[ERR] 需要 openpyxl")
        return 1

    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True)]

    records = []
    for r in rows:
        part = str(r[13]).strip() if r[13] else None   # N 接触网器材
        comp = str(r[14]).strip() if r[14] else None   # O 零部件
        desc = str(r[15]).strip() if r[15] else None   # P 缺陷描述
        if not (part and comp and desc):
            continue
        # 状态 = 描述 去掉器材前缀、去掉零部件
        state = desc
        if state.startswith(part):
            state = state[len(part):]
        if comp and state.startswith(comp):
            state = state[len(comp):]
        state = state.strip() or desc
        records.append({"part": part, "comp": comp, "desc": desc, "state": state})

    total = len(records)
    print("=" * 100)
    print(f"真实线路缺陷记录：{total} 条")
    print("=" * 100)

    # ---------- 器材层覆盖 ----------
    part_hist = collections.Counter(r["part"] for r in records)
    part_cov = {}
    covered_n = 0
    for p, n in part_hist.most_common():
        code, canon = match_legacy_part(p)
        part_cov[p] = {"n": n, "legacy": code, "legacy_canon": canon}
        if code:
            covered_n += n
    print(f"\n【器材层】现有 16 大类器材词能匹配的真实记录：{covered_n}/{total} = {covered_n/total:.1%}")

    uncovered_parts = [(p, v["n"]) for p, v in part_cov.items() if not v["legacy"]]
    uncovered_parts.sort(key=lambda x: -x[1])
    uncovered_n = sum(n for _, n in uncovered_parts)
    print(f"          未覆盖器材：{len(uncovered_parts)} 个，合计 {uncovered_n} 条 = {uncovered_n/total:.1%}")
    print(f"\n  未覆盖器材 TOP 20（按样本量）：")
    for p, n in uncovered_parts[:20]:
        print(f"    {n:>4} 条 ({n/total:>5.1%})  {p}")

    # ---------- 状态层覆盖 ----------
    st_hist = collections.Counter(r["state"] for r in records)
    st_cov = {}
    state_covered_n = 0
    for s, n in st_hist.most_common():
        legacy = match_legacy_state(s)
        st_cov[s] = {"n": n, "legacy": legacy}
        if legacy:
            state_covered_n += n
    print(f"\n【状态层】现有 16 大类状态原语能匹配的真实记录：{state_covered_n}/{total} = {state_covered_n/total:.1%}")

    uncovered_states = [(s, v["n"]) for s, v in st_cov.items() if not v["legacy"]]
    uncovered_states.sort(key=lambda x: -x[1])
    print(f"          未覆盖状态：{len(uncovered_states)} 种，合计 {total - state_covered_n} 条")
    print(f"\n  全部状态（TOP 20，✓=16大类已覆盖）：")
    for s, n in st_hist.most_common(20):
        mark = "✓" if st_cov[s]["legacy"] else "✗ 未覆盖"
        lg = st_cov[s]["legacy"] or ""
        print(f"    {n:>4} 条 ({n/total:>5.1%})  {s:<12} {mark} {lg}")

    # ---------- 联合覆盖（器材 AND 状态 都能匹配才算覆盖） ----------
    both = 0
    for r in records:
        pc, _ = match_legacy_part(r["part"])
        sc = match_legacy_state(r["state"])
        if pc and sc:
            both += 1
    print(f"\n【联合】器材 AND 状态 都能映射到现有 16 大类：{both}/{total} = {both/total:.1%}")
    print(f"        → 现有 16 大类对真实线路需求的覆盖率 ≈ {both/total:.1%}")
    print(f"        → 缺口 {total-both} 条 = {1-both/total:.1%}")

    # ---------- 零部件层（路径 2.5 部件头） ----------
    comp_hist = collections.Counter(r["comp"] for r in records)
    print(f"\n【零部件层】真实登记共 {len(comp_hist)} 种零部件（= 路径 2.5 部件头的类别数）：")
    for c, n in comp_hist.most_common():
        print(f"    {n:>4} 条 ({n/total:>5.1%})  {c}")

    # ---------- 持久化 ----------
    payload = {
        "total": total,
        "part_coverage": {"covered": covered_n, "ratio": covered_n / total,
                          "uncovered_parts": [{"part": p, "n": n} for p, n in uncovered_parts]},
        "state_coverage": {"covered": state_covered_n, "ratio": state_covered_n / total,
                           "uncovered_states": [{"state": s, "n": n} for s, n in uncovered_states]},
        "joint_coverage": {"covered": both, "ratio": both / total},
        "components": dict(comp_hist.most_common()),
        "part_hist": dict(part_hist.most_common()),
        "state_hist": dict(st_hist.most_common()),
    }
    op = OUT_DIR / "coverage_gap_analysis.json"
    op.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OK] 已保存 {op}")

    # ---------- 生成甲方待确认 Excel 模板 ----------
    try:
        from openpyxl import Workbook
        wb2 = Workbook()
        ws2 = wb2.active
        ws2.title = "器材映射待确认"
        ws2.append(["真实器材", "样本量", "占比", "是否现有16大类覆盖", "映射到的大类",
                    "建议归属(路径2.5部件头)", "甲方确认", "备注"])
        for p, n in part_hist.most_common():
            code, canon = match_legacy_part(p)
            ws2.append([p, n, f"{n/total:.1%}", "是" if code else "否", code or "",
                        "", "", "" if code else "需新增为部件头类别或说明"])
        ws3 = wb2.create_sheet("状态映射待确认")
        ws3.append(["真实状态", "样本量", "占比", "是否现有16大类覆盖", "映射到的状态原语",
                    "建议归属(路径2.5状态头10原语)", "甲方确认", "备注"])
        for s, n in st_hist.most_common():
            lg = match_legacy_state(s)
            ws3.append([s, n, f"{n/total:.1%}", "是" if lg else "否", lg or "", "", "",
                        "" if lg else "需新增状态原语"])
        ws4 = wb2.create_sheet("零部件清单")
        ws4.append(["零部件", "样本量", "占比", "说明"])
        for c, n in comp_hist.most_common():
            ws4.append([c, n, f"{n/total:.1%}", "路径2.5 部件头候选类别"])
        ox = OUT_DIR / "器材状态映射待甲方确认_20260901.xlsx"
        wb2.save(ox)
        print(f"[OK] 已生成甲方待确认模板 {ox}")
    except Exception as e:
        print(f"[WARN] Excel 模板生成失败: {e}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
